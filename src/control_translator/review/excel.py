"""Excel-based review workflow — export pending decisions for human sign-off.

Usage:
    ct export-review --config config\nzism-myjob.json --output review.xlsx
    # open in Excel, fill Decision column, save
    ct import-review --config config\nzism-myjob.json --input review.xlsx

The exported workbook has three sheets:

  Pending Review      — controls awaiting include/ignore decision
  OOS Candidates      — policies the LLM flagged as globally out of scope
  Preview Excluded    — informational: [Preview]: policies auto-filtered this run

The authority (e.g. GCSB/NCSC) fills in the Decision columns in Excel.
Import reads those decisions and updates the mapping store and OOS register.
"""
from __future__ import annotations

import json
import os
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── colour palette ─────────────────────────────────────────────────────────────
_NAVY  = "0E2D5A"
_BLUE  = "0078D4"
_WHITE = "FFFFFF"
_LIGHT = "EFF6FF"
_AMBER = "FFF4EC"
_GREEN = "EDF7ED"
_MID   = "C7DCEF"
_GREY  = "F8FAFD"


def _hdr_font(size=11):  return Font(name="Segoe UI", bold=True, color=_WHITE, size=size)
def _body_font(size=10): return Font(name="Segoe UI", size=size)
def _fill(hex_col):      return PatternFill("solid", fgColor=hex_col)
def _thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_row(ws, cols: list[str], row: int = 1, fill_hex: str = _NAVY) -> None:
    for col_idx, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font   = _hdr_font()
        cell.fill   = _fill(fill_hex)
        cell.border = _thin_border()
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _data_cell(ws, row, col, value, fill_hex=None, wrap=False, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font   = Font(name="Segoe UI", size=10, bold=bold)
    cell.border = _thin_border()
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")
    if fill_hex:
        cell.fill = _fill(fill_hex)
    return cell


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_dropdown(ws, col_letter: str, start_row: int, end_row: int,
                  options: list[str], prompt: str = "") -> None:
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                        showDropDown=False, showErrorMessage=True,
                        errorTitle="Invalid", error="Choose from the list.")
    if prompt:
        dv.prompt = prompt
        dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"


# ── export ────────────────────────────────────────────────────────────────────

def export_review(pipeline_result, *, output_path: str, framework_id: str,
                  version: str, oos_register_path: str | None = None,
                  oos_candidates: list | None = None,
                  preview_excluded: list | None = None) -> None:
    """Write the review workbook from a PipelineResult.

    oos_candidates and preview_excluded can be supplied explicitly when the
    pipeline result has empty lists due to carry-forward (no new classification).
    """
    if not HAS_OPENPYXL:
        raise SystemExit("openpyxl is required: pip install openpyxl")

    from ..models import Decision

    wb = openpyxl.Workbook()

    mapping  = pipeline_result.mapping
    catalog  = pipeline_result.catalog
    bundle   = pipeline_result.bundle

    ctrl_map = {c.id: c for c in catalog.controls()}

    # Resolve OOS candidates and preview excluded from args or bundle
    if oos_candidates is None:
        oos_candidates = (json.loads(bundle.files["oos-candidates.json"])
                         if bundle and "oos-candidates.json" in bundle.files else [])
    if preview_excluded is None:
        preview_excluded = mapping.preview_excluded or []

    # ── Sheet 1: Pending Review ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Pending Review"

    cols1 = ["Control ID", "Chapter", "Compliance", "Control Text",
             "Proposed Policy", "Policy GUID", "Confidence", "LLM Rationale",
             "Decision"]
    _header_row(ws1, cols1, row=1)
    ws1.row_dimensions[1].height = 22

    pending = sorted(mapping.pending_review(), key=lambda m: m.control_id)

    for r_idx, m in enumerate(pending, start=2):
        ctrl  = ctrl_map.get(m.control_id)
        prose = ctrl.prose[:300] if ctrl else ""
        chapter = (ctrl.family or "").split(".")[0].strip() if ctrl else ""
        compliance = (ctrl.props.get("compliance") or "") if ctrl else ""
        fill = _LIGHT if r_idx % 2 == 0 else None

        for policy in m.policies[:1]:  # primary policy per control
            _data_cell(ws1, r_idx, 1, m.control_id,    fill)
            _data_cell(ws1, r_idx, 2, chapter,          fill)
            _data_cell(ws1, r_idx, 3, compliance,       fill)
            _data_cell(ws1, r_idx, 4, prose,            fill, wrap=True)
            _data_cell(ws1, r_idx, 5, policy.display_name, fill)
            _data_cell(ws1, r_idx, 6, policy.policy_id.split("/")[-1], fill)
            _data_cell(ws1, r_idx, 7, round(m.confidence, 2), fill)
            _data_cell(ws1, r_idx, 8, m.rationale[:400], fill, wrap=True)
            # Decision cell — highlighted, empty for reviewer
            dec = ws1.cell(row=r_idx, column=9, value="")
            dec.fill   = _fill("FFFDE7")
            dec.font   = Font(name="Segoe UI", size=10, bold=True)
            dec.border = _thin_border()
            dec.alignment = Alignment(horizontal="center", vertical="center")

    if pending:
        _add_dropdown(ws1, "I", 2, len(pending) + 1,
                      ["Include", "Ignore", "Skip"],
                      "Include = approve mapping  |  Ignore = no policy maps this control  |  Skip = defer")

    _set_col_widths(ws1, [14, 8, 10, 50, 50, 36, 10, 50, 10])
    ws1.freeze_panes = "A2"

    # ── Sheet 2: OOS Candidates ───────────────────────────────────────────────
    # Columns: Policy Name | GUID | First Seen Control | Chapter | Compliance |
    #          Control Text | LLM OOS Reason | Action
    # Action dropdown includes "Include for this control" so a reviewer can
    # override an OOS flag and force the policy back into the initiative.
    ws2 = wb.create_sheet("OOS Candidates")
    cols2 = ["Policy Name", "Policy GUID", "First Seen Control",
             "Chapter", "Compliance", "Control Text",
             "Flagged by N controls", "All Flagging Controls",
             "LLM OOS Reason", "Action"]
    _header_row(ws2, cols2, row=1, fill_hex=_BLUE)
    ws2.row_dimensions[1].height = 22

    for r_idx, cand in enumerate(oos_candidates, start=2):
        fill = _AMBER if r_idx % 2 == 0 else "FFF9F0"
        ctrl_id = cand.get("first_seen_control", "")
        ctrl    = ctrl_map.get(ctrl_id)
        chapter    = (ctrl.family or "").split(".")[0].strip() if ctrl else ""
        compliance = (ctrl.props.get("compliance") or "")       if ctrl else ""
        prose      = (ctrl.prose or "")[:250]                   if ctrl else ""

        all_flagging = cand.get("flagging_controls", [ctrl_id] if ctrl_id else [])
        n_flagging   = cand.get("flagged_by_n_controls", len(all_flagging))

        _data_cell(ws2, r_idx, 1, cand.get("display_name", ""),        fill)
        _data_cell(ws2, r_idx, 2, cand.get("policy_id", "").split("/")[-1], fill)
        _data_cell(ws2, r_idx, 3, ctrl_id,                              fill)
        _data_cell(ws2, r_idx, 4, chapter,                              fill)
        _data_cell(ws2, r_idx, 5, compliance,                           fill)
        _data_cell(ws2, r_idx, 6, prose,                                fill, wrap=True)
        # show count prominently; if >1 list them all
        _data_cell(ws2, r_idx, 7, n_flagging,                           fill, bold=(n_flagging > 1))
        _data_cell(ws2, r_idx, 8, ", ".join(all_flagging[:20]),         fill, wrap=True)
        _data_cell(ws2, r_idx, 9, cand.get("oos_reason", ""),           fill, wrap=True)
        act = ws2.cell(row=r_idx, column=10, value="")
        act.fill      = _fill("FFFDE7")
        act.font      = Font(name="Segoe UI", size=10, bold=True)
        act.border    = _thin_border()
        act.alignment = Alignment(horizontal="center", vertical="center")

    if oos_candidates:
        _add_dropdown(
            ws2, "J", 2, len(oos_candidates) + 1,
            ["Add to nzism-ignore", "Global ignore",
             "Include for this control", "Skip"],
            "Add to nzism-ignore = NZISM exclusion  |  "
            "Global ignore = all standards  |  "
            "Include for this control = override OOS, keep in initiative")

    _set_col_widths(ws2, [55, 36, 16, 10, 11, 40, 8, 35, 50, 22])
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Preview Excluded (informational) ─────────────────────────────
    ws3 = wb.create_sheet("Preview Excluded (info)")
    cols3 = ["Policy Name", "Policy GUID", "Note"]
    _header_row(ws3, cols3, row=1, fill_hex="475569")
    for r_idx, prev in enumerate(preview_excluded, start=2):
        fill = _GREY if r_idx % 2 == 0 else None
        _data_cell(ws3, r_idx, 1, prev.get("display_name",""), fill)
        _data_cell(ws3, r_idx, 2, prev.get("policy_id","").split("/")[-1], fill)
        _data_cell(ws3, r_idx, 3, "Auto-excluded — will be reconsidered when GA", fill)
    _set_col_widths(ws3, [60, 36, 45])

    # ── Cover metadata ────────────────────────────────────────────────────────
    ws1["A1"].comment = None  # suppress any stale comments
    wb.properties.title   = f"Control Translator Review — {framework_id} v{version}"
    wb.properties.creator = "Control Translator"

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)


# ── import ────────────────────────────────────────────────────────────────────

def import_review(input_path: str, *, mapping_store_path: str,
                  oos_register_paths: list[str] | None = None,
                  corrections_path: str | None = None) -> dict:
    """Read decisions from the review workbook and update the mapping store.

    Returns a summary dict: {include, ignore, skipped, oos_added, oos_global}.
    """
    if not HAS_OPENPYXL:
        raise SystemExit("openpyxl is required: pip install openpyxl")

    from ..mapping.store import MappingStore, _norm_id
    from ..models import Decision

    wb   = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    summary = {"include": 0, "ignore": 0, "skipped": 0,
               "oos_added": 0, "oos_global": 0}

    # ── Sheet 1: Pending Review ───────────────────────────────────────────────
    store = MappingStore(mapping_store_path)
    mapping_set = store.load("", "")
    ws1 = wb["Pending Review"]
    rows1 = list(ws1.iter_rows(values_only=True))
    for row in rows1[1:]:
        if not row[0]: continue
        control_id = str(row[0]).strip()
        decision   = str(row[8]).strip().lower() if row[8] else ""
        if control_id not in mapping_set.mappings:
            continue
        if decision == "include":
            mapping_set.mappings[control_id].decision = Decision.INCLUDE
            summary["include"] += 1
        elif decision == "ignore":
            mapping_set.mappings[control_id].decision = Decision.IGNORE
            summary["ignore"] += 1
        else:
            summary["skipped"] += 1

    # ── Sheet 2: OOS Candidates ───────────────────────────────────────────────
    # Column layout (0-based): 0=Policy Name, 1=GUID, 2=First Seen Control,
    #   3=Chapter, 4=Compliance, 5=Control Text, 6=LLM OOS Reason, 7=Action
    if "OOS Candidates" in wb.sheetnames:
        ws2   = wb["OOS Candidates"]
        rows2 = list(ws2.iter_rows(values_only=True))
        nzism_path  = next((p for p in (oos_register_paths or [])
                            if "nzism" in p.lower()), None)
        global_path = next((p for p in (oos_register_paths or [])
                            if "global" in p.lower()
                            and "nzism" not in p.lower()), None)

        nzism_reg  = json.load(open(nzism_path,  encoding="utf-8")) if nzism_path  and os.path.exists(nzism_path)  else []
        global_reg = json.load(open(global_path, encoding="utf-8")) if global_path and os.path.exists(global_path) else []
        existing_guids = {_norm_id(e["policy_id"]) for e in nzism_reg + global_reg}

        for row in rows2[1:]:
            if not row[0]: continue
            action = str(row[9]).strip().lower() if len(row) > 9 and row[9] else ""
            if "skip" in action or not action: continue

            guid     = str(row[1]).strip() if row[1] else ""
            full_pid = f"/providers/Microsoft.Authorization/policyDefinitions/{guid.lower()}"
            policy_name = str(row[0]).strip() if row[0] else ""
            oos_reason  = str(row[6]).strip() if len(row) > 6 and row[6] else ""

            if "include for this control" in action:
                # Override: mark the first_seen_control as INCLUDE with this policy
                ctrl_id = str(row[2]).strip() if row[2] else ""
                if ctrl_id and ctrl_id in mapping_set.mappings:
                    from ..models import Decision, ControlMapping, PolicyRef
                    existing_m = mapping_set.mappings[ctrl_id]
                    if existing_m.decision != Decision.IGNORE:
                        mapping_set.mappings[ctrl_id] = ControlMapping(
                            control_id=ctrl_id,
                            decision=Decision.INCLUDE,
                            policies=[PolicyRef(policy_id=full_pid,
                                                display_name=policy_name)],
                            rationale=f"Reviewer override — included despite OOS flag: {oos_reason[:200]}",
                            source="review-override",
                            confidence=1.0,
                        )
                        summary["include"] += 1

                # Write to corrections file for future few-shot learning
                if corrections_path and ctrl_id:
                    chapter  = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                    compliance = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                    from ..mapping.corrections import save_correction
                    save_correction(corrections_path, {
                        "policy_id":         full_pid,
                        "display_name":      policy_name,
                        "control_id":        ctrl_id,
                        "chapter":           chapter,
                        "compliance":        compliance,
                        "include_reasoning": (f"Reviewer confirmed relevant for {chapter} "
                                              f"controls despite OOS flag: {oos_reason[:180]}"),
                        "added_date":        str(date.today()),
                        "source":            "review-override",
                    })
                continue  # don't add to any ignore file

            if _norm_id(full_pid) in existing_guids:
                continue
            entry = {
                "policy_id":    full_pid,
                "display_name": policy_name,
                "reason":       oos_reason or "Flagged by AI classifier.",
                "oos_date":     str(date.today()),
                "source":       "review-import",
            }
            if "global" in action:
                global_reg.append(entry)
                summary["oos_global"] += 1
            else:
                nzism_reg.append(entry)
                summary["oos_added"] += 1

        if nzism_path:
            json.dump(nzism_reg, open(nzism_path, "w", encoding="utf-8"),
                      indent=2, ensure_ascii=False)
        if global_path is not None:
            json.dump(global_reg, open(global_path, "w", encoding="utf-8"),
                      indent=2, ensure_ascii=False)

    store.save(mapping_set)

    return summary
